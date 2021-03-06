
# Python program to illustrate a stop watch
# using Tkinter
#importing the required libraries
import tkinter as Tkinter
from openpyxl import load_workbook
from datetime import datetime
from datetime import date
import xlsxwriter

from sqlalchemy import column

counter = 0
running = False
def counter_label(label):
    def count():
        if running:
            global counter
   
            # To manage the initial delay.
            if counter==0:            
                display="Starting..."
            else:
                tt = datetime.fromtimestamp(counter)
                string = tt.strftime("%H:%M:%S")
                display=string
   
            label['text']=display   # Or label.config(text=display)
   
            # label.after(arg1, arg2) delays by 
            # first argument given in milliseconds
            # and then calls the function given as second argument.
            # Generally like here we need to call the 
            # function in which it is present repeatedly.
            # Delays by 1000ms=1 seconds and call count again.
            label.after(1000, count) 
            counter += 1
   
    # Triggering the start of the counter.
    count()     
   
# start function of the stopwatch
def Start(label):
    global running
    running=True
    counter_label(label)
    start['state']='disabled'
    stop['state']='normal'
    reset['state']='normal'
    save['state']='disabled'
   
# Stop function of the stopwatch
def Stop():
    global running
    start['state']='normal'
    stop['state']='disabled'
    reset['state']='normal'
    save['state']='normal'
    running = False

def Save(label):
    global running
    start['state']='normal'
    stop['state']='disabled'
    reset['state']='normal' 
    save['state']="disabled"
    label['text']='Saved'
    Saving()


# Reset function of the stopwatch
def Reset(label):
    global counter
    counter=0
   
    # If rest is pressed after pressing stop.
    if running==False:      
        reset['state']='disabled'
        label['text']='Welcome!'
   
    # If reset is pressed while the stopwatch is running.
    else:               
        label['text']='Starting...'

def Saving():
    writen = False
    row = 1
    col = 1

    wb = load_workbook("C:/Users/joese/Desktop/Work.xlsx")
    ws = wb['Sheet1']
    dayNum = date.today().strftime("%B %d, %Y") 
    string = datetime.fromtimestamp(counter).strftime("%H:%M:%S")
    while writen == False:
        if ws.cell(row, col).value == None: 
            ws.cell(row, col).value = dayNum  	
            ws.cell(row, col+1).value = string
            wb.save("C:/Users/joese/Desktop/work.xlsx")
            writen=True 
        row += 1



root = Tkinter.Tk()
root.title("Stopwatch")
   
# Fixing the window size.
root.minsize(width=250, height=70)
label = Tkinter.Label(root, text="Welcome!", fg="black", font="Verdana 30 bold")
label.pack()
f = Tkinter.Frame(root)
start = Tkinter.Button(f, text='Start', width=6, command=lambda:Start(label))
stop = Tkinter.Button(f, text='Stop',width=6,state='disabled', command=Stop)
reset = Tkinter.Button(f, text='Reset',width=6, state='disabled', command=lambda:Reset(label))
save = Tkinter.Button(f, text = "Save", width=6, state='disabled', command=lambda:Save(label))
f.pack(anchor = 'center',pady=5)
start.pack(side="left")
stop.pack(side ="left")
reset.pack(side="left")
save.pack(side = "left")
root.mainloop()